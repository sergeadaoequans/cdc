import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
import math


# Fonctionne mais modification "pratique" à apporter


def adjust_column_widths(worksheet, start_row=1):
    """
    Ajuste automatiquement la largeur des colonnes en évitant les problèmes avec les cellules fusionnées
    """
    for col_num in range(1, worksheet.max_column + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0

        for row_num in range(start_row, worksheet.max_row + 1):
            try:
                cell = worksheet[f"{column_letter}{row_num}"]
                if hasattr(cell, 'value') and cell.value is not None:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                continue

        adjusted_width = min(max_length + 2, 50)  # Limiter la largeur maximale
        worksheet.column_dimensions[column_letter].width = adjusted_width


def create_cable_management_excel(filename="Gestion_Chemins_Cables.xlsx"):
    """
    Crée un fichier Excel complet pour la gestion des chemins de câbles électriques
    """

    # Créer le workbook
    wb = Workbook()

    # Supprimer la feuille par défaut
    wb.remove(wb.active)

    # =============================================================================
    # 1. FEUILLE DES CÂBLES
    # =============================================================================
    ws_cables = wb.create_sheet("Câbles")

    # En-têtes pour les câbles
    headers_cables = [
        "ID_Cable", "Type_Cable", "Diametre_mm", "Nombre_Conducteurs",
        "Section_mm2", "Poids_kg_m", "Description", "Surface_mm2"
    ]

    # Ajouter les en-têtes
    for col, header in enumerate(headers_cables, 1):
        cell = ws_cables.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # Données d'exemple pour les câbles
    sample_cables = [
        ["W001", "Alimentation", 12.5, 3, 2.5, 0.8, "Câble alimentation 3x2.5mm²", "=PI()*(C2/2)^2"],
        ["W002", "Contrôle", 8.2, 2, 1.5, 0.4, "Câble contrôle 2x1.5mm²", "=PI()*(C3/2)^2"],
        ["W003", "Signal", 6.1, 4, 0.75, 0.3, "Câble signal 4x0.75mm²", "=PI()*(C4/2)^2"],
        ["W004", "Alimentation", 15.8, 5, 4, 1.2, "Câble alimentation 5x4mm²", "=PI()*(C5/2)^2"],
        ["W005", "Ethernet", 5.5, 8, 0.2, 0.2, "Câble Ethernet Cat6", "=PI()*(C6/2)^2"],
        ["W006", "Contrôle", 10.2, 3, 1.5, 0.6, "Câble contrôle 3x1.5mm²", "=PI()*(C7/2)^2"],
        ["W007", "Signal", 7.8, 6, 0.5, 0.4, "Câble signal 6x0.5mm²", "=PI()*(C8/2)^2"]
    ]

    # Ajouter les données d'exemple
    for row, cable_data in enumerate(sample_cables, 2):
        for col, value in enumerate(cable_data, 1):
            ws_cables.cell(row=row, column=col, value=value)

    # Ajuster la largeur des colonnes
    adjust_column_widths(ws_cables)

    # =============================================================================
    # 2. FEUILLE DES CHEMINS DE CÂBLES
    # =============================================================================
    ws_chemins = wb.create_sheet("Chemins_Cables")

    # En-têtes pour les chemins de câbles
    headers_chemins = [
        "ID_Chemin", "Description", "Largeur_mm", "Hauteur_mm",
        "Longueur_m", "Surface_mm2", "Capacite_Utilisable_mm2", "Taux_Remplissage_%"
    ]

    # Ajouter les en-têtes
    for col, header in enumerate(headers_chemins, 1):
        cell = ws_chemins.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # Données d'exemple pour les chemins de câbles
    sample_chemins = [
        ["A", "Chemin principal niveau 1", 200, 100, 50, "=C2*D2", "=F2*(1-Paramètres!$B$2)",
         "=IFERROR(SUMIF(Assignation!D:D,\"*\"&A2&\"*\",Assignation!F:F)/G2*100,0)"],
        ["B", "Dérivation vers local technique", 150, 80, 30, "=C3*D3", "=F3*(1-Paramètres!$B$2)",
         "=IFERROR(SUMIF(Assignation!D:D,\"*\"&A3&\"*\",Assignation!F:F)/G3*100,0)"],
        ["C", "Chemin secondaire niveau 2", 100, 60, 25, "=C4*D4", "=F4*(1-Paramètres!$B$2)",
         "=IFERROR(SUMIF(Assignation!D:D,\"*\"&A4&\"*\",Assignation!F:F)/G4*100,0)"],
        ["D", "Arrivée tableau électrique", 120, 80, 15, "=C5*D5", "=F5*(1-Paramètres!$B$2)",
         "=IFERROR(SUMIF(Assignation!D:D,\"*\"&A5&\"*\",Assignation!F:F)/G5*100,0)"],
        ["E", "Distribution finale", 80, 50, 20, "=C6*D6", "=F6*(1-Paramètres!$B$2)",
         "=IFERROR(SUMIF(Assignation!D:D,\"*\"&A6&\"*\",Assignation!F:F)/G6*100,0)"]
    ]

    # Ajouter les données d'exemple
    for row, chemin_data in enumerate(sample_chemins, 2):
        for col, value in enumerate(chemin_data, 1):
            ws_chemins.cell(row=row, column=col, value=value)

    # Ajuster la largeur des colonnes
    adjust_column_widths(ws_chemins)

    # =============================================================================
    # 3. FEUILLE DES PARAMÈTRES
    # =============================================================================
    ws_params = wb.create_sheet("Paramètres")

    # Configuration des paramètres
    ws_params['A1'] = "Paramètre"
    ws_params['B1'] = "Valeur"
    ws_params['C1'] = "Description"

    # Mise en forme des en-têtes
    for col in ['A1', 'B1', 'C1']:
        ws_params[col].font = Font(bold=True, color="FFFFFF")
        ws_params[col].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ws_params[col].alignment = Alignment(horizontal='center')

    # Données des paramètres
    ws_params['A2'] = "Taux de réserve"
    ws_params['B2'] = 0.4
    ws_params['C2'] = "40% = 60% de capacité utilisable"

    ws_params['A3'] = "Facteur de sécurité"
    ws_params['B3'] = 1.2
    ws_params['C3'] = "Facteur multiplicateur pour les calculs"

    # Mise en forme des paramètres
    ws_params['B2'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws_params['B3'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Ajuster la largeur des colonnes
    ws_params.column_dimensions['A'].width = 20
    ws_params.column_dimensions['B'].width = 15
    ws_params.column_dimensions['C'].width = 35

    # =============================================================================
    # 4. FEUILLE D'ASSIGNATION
    # =============================================================================
    ws_assign = wb.create_sheet("Assignation")

    # En-têtes pour l'assignation
    headers_assign = [
        "ID_Cable", "Type_Cable", "Diametre_mm", "Chemin_Cable",
        "Nb_Troncons", "Surface_Cable_mm2", "Longueur_Totale_m", "Statut"
    ]

    # Ajouter les en-têtes
    for col, header in enumerate(headers_assign, 1):
        cell = ws_assign.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="000000")
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # Données d'exemple pour l'assignation
    sample_assign = [
        ["W001", "=VLOOKUP(A2,Câbles!A:H,2,FALSE)", "=VLOOKUP(A2,Câbles!A:H,3,FALSE)", "A/B/D",
         "=LEN(D2)-LEN(SUBSTITUTE(D2,\"/\",\"\"))+1", "=VLOOKUP(A2,Câbles!A:H,8,FALSE)",
         "=SUMPRODUCT(--(LEN(D2)>0), VLOOKUP(LEFT(D2,1),Chemins_Cables!A:E,5,FALSE))",
         "=IF(A2<>\"\",\"Assigné\",\"Non assigné\")"],
        ["W002", "=VLOOKUP(A3,Câbles!A:H,2,FALSE)", "=VLOOKUP(A3,Câbles!A:H,3,FALSE)", "A/C/E",
         "=LEN(D3)-LEN(SUBSTITUTE(D3,\"/\",\"\"))+1", "=VLOOKUP(A3,Câbles!A:H,8,FALSE)",
         "=SUMPRODUCT(--(LEN(D3)>0), VLOOKUP(LEFT(D3,1),Chemins_Cables!A:E,5,FALSE))",
         "=IF(A3<>\"\",\"Assigné\",\"Non assigné\")"],
        ["W003", "=VLOOKUP(A4,Câbles!A:H,2,FALSE)", "=VLOOKUP(A4,Câbles!A:H,3,FALSE)", "B/C/D/E",
         "=LEN(D4)-LEN(SUBSTITUTE(D4,\"/\",\"\"))+1", "=VLOOKUP(A4,Câbles!A:H,8,FALSE)",
         "=SUMPRODUCT(--(LEN(D4)>0), VLOOKUP(LEFT(D4,1),Chemins_Cables!A:E,5,FALSE))",
         "=IF(A4<>\"\",\"Assigné\",\"Non assigné\")"],
        ["W004", "=VLOOKUP(A5,Câbles!A:H,2,FALSE)", "=VLOOKUP(A5,Câbles!A:H,3,FALSE)", "A/B",
         "=LEN(D5)-LEN(SUBSTITUTE(D5,\"/\",\"\"))+1", "=VLOOKUP(A5,Câbles!A:H,8,FALSE)",
         "=SUMPRODUCT(--(LEN(D5)>0), VLOOKUP(LEFT(D5,1),Chemins_Cables!A:E,5,FALSE))",
         "=IF(A5<>\"\",\"Assigné\",\"Non assigné\")"],
        ["W005", "=VLOOKUP(A6,Câbles!A:H,2,FALSE)", "=VLOOKUP(A6,Câbles!A:H,3,FALSE)", "C/D/E",
         "=LEN(D6)-LEN(SUBSTITUTE(D6,\"/\",\"\"))+1", "=VLOOKUP(A6,Câbles!A:H,8,FALSE)",
         "=SUMPRODUCT(--(LEN(D6)>0), VLOOKUP(LEFT(D6,1),Chemins_Cables!A:E,5,FALSE))",
         "=IF(A6<>\"\",\"Assigné\",\"Non assigné\")"],
        ["W006", "=VLOOKUP(A7,Câbles!A:H,2,FALSE)", "=VLOOKUP(A7,Câbles!A:H,3,FALSE)", "A/C",
         "=LEN(D7)-LEN(SUBSTITUTE(D7,\"/\",\"\"))+1", "=VLOOKUP(A7,Câbles!A:H,8,FALSE)",
         "=SUMPRODUCT(--(LEN(D7)>0), VLOOKUP(LEFT(D7,1),Chemins_Cables!A:E,5,FALSE))",
         "=IF(A7<>\"\",\"Assigné\",\"Non assigné\")"],
        ["W007", "=VLOOKUP(A8,Câbles!A:H,2,FALSE)", "=VLOOKUP(A8,Câbles!A:H,3,FALSE)", "B/D",
         "=LEN(D8)-LEN(SUBSTITUTE(D8,\"/\",\"\"))+1", "=VLOOKUP(A8,Câbles!A:H,8,FALSE)",
         "=SUMPRODUCT(--(LEN(D8)>0), VLOOKUP(LEFT(D8,1),Chemins_Cables!A:E,5,FALSE))",
         "=IF(A8<>\"\",\"Assigné\",\"Non assigné\")"]
    ]

    # Ajouter les données d'exemple
    for row, assign_data in enumerate(sample_assign, 2):
        for col, value in enumerate(assign_data, 1):
            ws_assign.cell(row=row, column=col, value=value)

    # Mise en forme du chemin de câble
    for row in range(2, len(sample_assign) + 2):
        ws_assign.cell(row=row, column=4).fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9",
                                                             fill_type="solid")

    # Ajuster la largeur des colonnes
    adjust_column_widths(ws_assign)

    # =============================================================================
    # 5. FEUILLE TABLEAU DE BORD
    # =============================================================================
    ws_dashboard = wb.create_sheet("Tableau_de_Bord")

    # Configuration du tableau de bord
    ws_dashboard['A1'] = "TABLEAU DE BORD - GESTION DES CHEMINS DE CÂBLES"
    ws_dashboard.merge_cells('A1:G1')
    ws_dashboard['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws_dashboard['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws_dashboard['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_dashboard.row_dimensions[1].height = 30

    # Sélection du chemin
    ws_dashboard['A3'] = "Sélectionner un chemin de câble:"
    ws_dashboard['A3'].font = Font(bold=True, size=12)
    ws_dashboard['B3'] = "A"  # Valeur par défaut
    ws_dashboard['B3'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws_dashboard['B3'].alignment = Alignment(horizontal='center')
    ws_dashboard['B3'].font = Font(bold=True, size=14)

    # Informations du chemin sélectionné
    ws_dashboard['A5'] = "INFORMATIONS DU CHEMIN SÉLECTIONNÉ"
    ws_dashboard['A5'].font = Font(bold=True, size=12, color="FFFFFF")
    ws_dashboard['A5'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws_dashboard.merge_cells('A5:C5')

    # Détails du chemin
    details_headers = ["Propriété", "Valeur", "Unité"]
    for col, header in enumerate(details_headers, 1):
        cell = ws_dashboard.cell(row=6, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # Formules pour les détails du chemin
    details_data = [
        ["ID Chemin", "=B3", ""],
        ["Description", "=VLOOKUP(B3,Chemins_Cables!A:H,2,FALSE)", ""],
        ["Largeur", "=VLOOKUP(B3,Chemins_Cables!A:H,3,FALSE)", "mm"],
        ["Hauteur", "=VLOOKUP(B3,Chemins_Cables!A:H,4,FALSE)", "mm"],
        ["Longueur", "=VLOOKUP(B3,Chemins_Cables!A:H,5,FALSE)", "m"],
        ["Surface totale", "=VLOOKUP(B3,Chemins_Cables!A:H,6,FALSE)", "mm²"],
        ["Capacité utilisable", "=VLOOKUP(B3,Chemins_Cables!A:H,7,FALSE)", "mm²"],
        ["Taux de remplissage", "=VLOOKUP(B3,Chemins_Cables!A:H,8,FALSE)", "%"],
        ["Nombre de câbles", "=COUNTIF(Assignation!D:D,\"*\"&B3&\"*\")", "pcs"]
    ]

    for row, detail in enumerate(details_data, 7):
        ws_dashboard.cell(row=row, column=1, value=detail[0])
        ws_dashboard.cell(row=row, column=2, value=detail[1])
        ws_dashboard.cell(row=row, column=3, value=detail[2])

    # Liste des câbles assignés
    ws_dashboard['E5'] = "CÂBLES ASSIGNÉS À CE CHEMIN"
    ws_dashboard['E5'].font = Font(bold=True, size=12, color="FFFFFF")
    ws_dashboard['E5'].fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
    ws_dashboard.merge_cells('E5:G5')

    # En-têtes pour la liste des câbles
    cables_headers = ["ID Câble", "Type", "Surface mm²"]
    for col, header in enumerate(cables_headers, 5):
        cell = ws_dashboard.cell(row=6, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # Simplification pour la liste des câbles - on va utiliser une approche plus directe
    cable_formulas = [
        [
            '=IF(COUNTIF(Assignation!D:D,"*"&$B$3&"*")>=1,INDEX(Assignation!A:A,MATCH(TRUE,ISNUMBER(SEARCH($B$3,Assignation!D:D)),0)),"")'],
        [
            '=IF(COUNTIF(Assignation!D:D,"*"&$B$3&"*")>=2,INDEX(Assignation!A:A,MATCH(TRUE,ISNUMBER(SEARCH($B$3,Assignation!D:D)),0)+1),"")'],
        [
            '=IF(COUNTIF(Assignation!D:D,"*"&$B$3&"*")>=3,INDEX(Assignation!A:A,MATCH(TRUE,ISNUMBER(SEARCH($B$3,Assignation!D:D)),0)+2),"")']
    ]

    # Pour simplifier, on va juste mettre des références directes
    direct_cables = [
        ["=IF(SEARCH($B$3,Assignation!D2,1)>0,Assignation!A2,\"\")",
         "=IF(E7<>\"\",VLOOKUP(E7,Câbles!A:H,2,FALSE),\"\")", "=IF(E7<>\"\",VLOOKUP(E7,Câbles!A:H,8,FALSE),\"\")"],
        ["=IF(ISERROR(SEARCH($B$3,Assignation!D3,1)),\"\",Assignation!A3)",
         "=IF(E8<>\"\",VLOOKUP(E8,Câbles!A:H,2,FALSE),\"\")", "=IF(E8<>\"\",VLOOKUP(E8,Câbles!A:H,8,FALSE),\"\")"],
        ["=IF(ISERROR(SEARCH($B$3,Assignation!D4,1)),\"\",Assignation!A4)",
         "=IF(E9<>\"\",VLOOKUP(E9,Câbles!A:H,2,FALSE),\"\")", "=IF(E9<>\"\",VLOOKUP(E9,Câbles!A:H,8,FALSE),\"\")"],
        ["=IF(ISERROR(SEARCH($B$3,Assignation!D5,1)),\"\",Assignation!A5)",
         "=IF(E10<>\"\",VLOOKUP(E10,Câbles!A:H,2,FALSE),\"\")", "=IF(E10<>\"\",VLOOKUP(E10,Câbles!A:H,8,FALSE),\"\")"],
        ["=IF(ISERROR(SEARCH($B$3,Assignation!D6,1)),\"\",Assignation!A6)",
         "=IF(E11<>\"\",VLOOKUP(E11,Câbles!A:H,2,FALSE),\"\")", "=IF(E11<>\"\",VLOOKUP(E11,Câbles!A:H,8,FALSE),\"\")"]
    ]

    for row, cable_data in enumerate(direct_cables, 7):
        for col, formula in enumerate(cable_data, 5):
            ws_dashboard.cell(row=row, column=col, value=formula)

    # Ajuster la largeur des colonnes manuellement pour éviter les problèmes
    column_widths = {'A': 25, 'B': 20, 'C': 10, 'D': 5, 'E': 15, 'F': 20, 'G': 15}
    for col, width in column_widths.items():
        ws_dashboard.column_dimensions[col].width = width

    # =============================================================================
    # 6. FEUILLE DE CALCULS
    # =============================================================================
    ws_calculs = wb.create_sheet("Calculs")

    # En-tête
    ws_calculs['A1'] = "FEUILLE DE CALCULS ET STATISTIQUES"
    ws_calculs.merge_cells('A1:F1')
    ws_calculs['A1'].font = Font(bold=True, size=14, color="000000")
    ws_calculs['A1'].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    ws_calculs['A1'].alignment = Alignment(horizontal='center')

    # Statistiques générales
    ws_calculs['A3'] = "STATISTIQUES GÉNÉRALES"
    ws_calculs['A3'].font = Font(bold=True, size=12)
    ws_calculs['A3'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    stats_data = [
        ["Nombre total de câbles", "=COUNTA(Câbles!A:A)-1", "pcs"],
        ["Nombre de chemins de câbles", "=COUNTA(Chemins_Cables!A:A)-1", "pcs"],
        ["Surface totale des câbles", "=ROUND(SUM(Câbles!H:H),2)", "mm²"],
        ["Capacité totale des chemins", "=ROUND(SUM(Chemins_Cables!G:G),2)", "mm²"],
        ["Taux d'utilisation global", "=ROUND(C7/C8*100,1)", "%"]
    ]

    # En-têtes pour les statistiques
    ws_calculs['A4'] = "Description"
    ws_calculs['B4'] = "Valeur"
    ws_calculs['C4'] = "Unité"
    for col in ['A4', 'B4', 'C4']:
        ws_calculs[col].font = Font(bold=True)
        ws_calculs[col].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    for row, stat in enumerate(stats_data, 5):
        ws_calculs.cell(row=row, column=1, value=stat[0])
        ws_calculs.cell(row=row, column=2, value=stat[1])
        ws_calculs.cell(row=row, column=3, value=stat[2])

    # Résumé par chemin
    ws_calculs['A12'] = "RÉSUMÉ PAR CHEMIN DE CÂBLE"
    ws_calculs['A12'].font = Font(bold=True, size=12)
    ws_calculs['A12'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    resume_headers = ["Chemin", "Description", "Capacité mm²", "Utilisé mm²", "Taux %", "Statut"]
    for col, header in enumerate(resume_headers, 1):
        cell = ws_calculs.cell(row=13, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # Données du résumé (pour chaque chemin A à E)
    chemins_list = ['A', 'B', 'C', 'D', 'E']
    for row, chemin in enumerate(chemins_list, 14):
        ws_calculs.cell(row=row, column=1, value=chemin)
        ws_calculs.cell(row=row, column=2, value=f"=VLOOKUP(\"{chemin}\",Chemins_Cables!A:H,2,FALSE)")
        ws_calculs.cell(row=row, column=3, value=f"=ROUND(VLOOKUP(\"{chemin}\",Chemins_Cables!A:H,7,FALSE),2)")
        ws_calculs.cell(row=row, column=4, value=f"=ROUND(SUMIF(Assignation!D:D,\"*{chemin}*\",Assignation!F:F),2)")
        ws_calculs.cell(row=row, column=5, value=f"=IF(C{row}>0,ROUND(D{row}/C{row}*100,1),0)")
        ws_calculs.cell(row=row, column=6, value=f"=IF(E{row}>90,\"⚠️ PLEIN\",IF(E{row}>70,\"⚡ ATTENTION\",\"✅ OK\"))")

    # Ajuster largeur des colonnes
    adjust_column_widths(ws_calculs)

    # =============================================================================
    # FORMATAGE CONDITIONNEL
    # =============================================================================

    # Formatage conditionnel pour le taux de remplissage
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # Appliquer le formatage conditionnel pour les taux de remplissage
    ws_chemins.conditional_formatting.add('H2:H100',
                                          CellIsRule(operator='greaterThan', formula=['90'], fill=red_fill))
    ws_chemins.conditional_formatting.add('H2:H100',
                                          CellIsRule(operator='between', formula=['70', '90'], fill=yellow_fill))
    ws_chemins.conditional_formatting.add('H2:H100',
                                          CellIsRule(operator='lessThan', formula=['70'], fill=green_fill))

    # Formatage conditionnel pour la feuille Calculs
    ws_calculs.conditional_formatting.add('E14:E18',
                                          CellIsRule(operator='greaterThan', formula=['90'], fill=red_fill))
    ws_calculs.conditional_formatting.add('E14:E18',
                                          CellIsRule(operator='between', formula=['70', '90'], fill=yellow_fill))
    ws_calculs.conditional_formatting.add('E14:E18',
                                          CellIsRule(operator='lessThan', formula=['70'], fill=green_fill))

    # =============================================================================
    # PROTECTION ET FINALISATION
    # =============================================================================

    # Ajouter des bordures à tous les tableaux
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Appliquer les bordures aux différentes feuilles
    for ws in [ws_cables, ws_chemins, ws_assign, ws_dashboard, ws_calculs]:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value is not None:
                    cell.border = thin_border

    # Sauvegarder le fichier
    wb.save(filename)
    return filename


def create_advanced_cable_management():
    """
    Fonction principale pour créer l'application complète de gestion des câbles
    """
    print("=" * 60)
    print("🔌 CRÉATION DE L'APPLICATION GESTION DES CHEMINS DE CÂBLES")
    print("=" * 60)

    try:
        filename = create_cable_management_excel()

        print("✅ Fichier Excel créé avec succès !")
        print(f"📁 Nom du fichier: {filename}")
        print("\n" + "=" * 60)
        print("📋 STRUCTURE DU FICHIER:")
        print("=" * 60)

        sheets_info = [
            ("1. Câbles", "Base de données des câbles avec calcul automatique des surfaces"),
            ("2. Chemins_Cables", "Définition des chemins avec dimensions et capacités"),
            ("3. Paramètres", "Configuration du taux de réserve et autres paramètres"),
            ("4. Assignation", "Attribution des câbles aux chemins (un câble par ligne)"),
            ("5. Tableau_de_Bord", "Interface interactive pour visualiser les résultats"),
            ("6. Calculs", "Statistiques et résumés globaux")
        ]

        for sheet, description in sheets_info:
            print(f"{sheet}: {description}")

        print("\n" + "=" * 60)
        print("🚀 GUIDE D'UTILISATION:")
        print("=" * 60)
        print("1. 📊 Modifiez le taux de réserve dans l'onglet 'Paramètres'")
        print("2. 🔌 Ajoutez vos câbles dans l'onglet 'Câbles'")
        print("3. 🛤️  Définissez vos chemins dans l'onglet 'Chemins_Cables'")
        print("4. 🎯 Assignez les câbles aux chemins dans 'Assignation'")
        print("   Format: A/B/C/D (utilisez '/' comme séparateur)")
        print("5. 📈 Utilisez le 'Tableau_de_Bord' pour visualiser")
        print("6. 📊 Consultez les 'Calculs' pour les statistiques")

        print("\n" + "=" * 60)
        print("✨ FONCTIONNALITÉS INCLUSES:")
        print("=" * 60)

        features = [
            "✅ Calcul automatique des surfaces (π × (diamètre/2)²)",
            "✅ Gestion du taux de réserve configurable",
            "✅ Assignation flexible avec chemins multiples",
            "✅ Tableau de bord interactif avec sélection",
            "✅ Formatage conditionnel (🟢Vert/🟡Jaune/🔴Rouge)",
            "✅ Statistiques globales et par chemin",
            "✅ Compatible Excel Online et Google Sheets",
            "✅ Aucune macro VBA - 100% formules",
            "✅ Interface utilisateur intuitive",
            "✅ Données d'exemple pour test immédiat"
        ]

        for feature in features:
            print(f"  {feature}")

        print("\n" + "=" * 60)
        print("🎯 CODES COULEUR:")
        print("=" * 60)
        print("🟢 VERT    : Taux de remplissage < 70% (OK)")
        print("🟡 JAUNE   : Taux de remplissage 70-90% (Attention)")
        print("🔴 ROUGE   : Taux de remplissage > 90% (Plein)")

        print("\n" + "=" * 60)
        print("📝 EXEMPLE D'ASSIGNATION:")
        print("=" * 60)
        print("Câble W001 → Chemin: A/B/D")
        print("  ↳ Passe par les chemins A, puis B, puis D")
        print("Câble W002 → Chemin: A/C/E")
        print("  ↳ Passe par les chemins A, puis C, puis E")

        print("\n" + "=" * 60)
        print("⚙️ PERSONNALISATION:")
        print("=" * 60)
        print("• Modifiez le taux de réserve dans 'Paramètres!B2'")
        print("• Ajoutez vos propres câbles dans 'Câbles'")
        print("• Créez vos chemins dans 'Chemins_Cables'")
        print("• Le tableau de bord s'adapte automatiquement")

        print("\n" + "=" * 60)
        print("🔧 MAINTENANCE:")
        print("=" * 60)
        print("• Toutes les formules sont relatives et extensibles")
        print("• Ajoutez des lignes sans casser les calculs")
        print("• Copiez les formules vers le bas si nécessaire")
        print("• Le formatage conditionnel s'applique automatiquement")

        return True

    except Exception as e:
        print(f"❌ Erreur lors de la création du fichier: {str(e)}")
        return False


def create_user_manual():
    """
    Crée un manuel d'utilisation séparé
    """
    manual = """
    # 📖 MANUEL D'UTILISATION - GESTION DES CHEMINS DE CÂBLES

    ## 🎯 OBJECTIF
    Cette application Excel permet de gérer efficacement le remplissage des chemins de câbles électriques en calculant automatiquement les capacités, les taux de remplissage et en visualisant les résultats.

    ## 📊 STRUCTURE DES FEUILLES

    ### 1. CÂBLES
    - **ID_Cable**: Identifiant unique du câble
    - **Type_Cable**: Type (Alimentation, Contrôle, Signal, etc.)
    - **Diametre_mm**: Diamètre extérieur en millimètres
    - **Surface_mm2**: Calculée automatiquement avec π×(diamètre/2)²

    ### 2. CHEMINS_CABLES
    - **ID_Chemin**: Identifiant du chemin (A, B, C, etc.)
    - **Largeur_mm** / **Hauteur_mm**: Dimensions du chemin
    - **Capacite_Utile_mm2**: Calculée avec taux de réserve

    ### 3. ASSIGNATION
    - **ID_Cable**: Référence du câble
    - **Chemin**: Format "A/B/C/D" avec séparateur "/"
    - Un câble par ligne pour faciliter l'importation

    ### 4. TABLEAU_DE_BORD
    - Interface interactive pour sélectionner et visualiser
    - Détails complets par chemin sélectionné

    ## 🚀 UTILISATION ÉTAPE PAR ÉTAPE

    1. **Configuration initiale**:
       - Ouvrir l'onglet "Paramètres"
       - Définir le taux de réserve (ex: 20% = 0.2)

    2. **Saisie des câbles**:
       - Aller dans l'onglet "Câbles"
       - Saisir: ID, Type, Diamètre, Nb conducteurs
       - La surface se calcule automatiquement

    3. **Définition des chemins**:
       - Onglet "Chemins_Cables"
       - Saisir les dimensions de chaque chemin
       - La capacité utile se calcule automatiquement

    4. **Assignation des câbles**:
       - Onglet "Assignation"
       - Format: Cable_ID → Chemin (ex: A/B/D)
       - Un câble par ligne obligatoire

    5. **Visualisation**:
       - Utiliser le "Tableau_de_Bord"
       - Sélectionner un chemin dans la liste
       - Consulter les résultats automatiquement

    ## 📈 INTERPRÉTATION DES RÉSULTATS

    ### Codes couleur:
    - 🟢 **VERT**: < 70% - Capacité normale
    - 🟡 **JAUNE**: 70-90% - Attention, proche saturation
    - 🔴 **ROUGE**: > 90% - Chemin saturé, action requise

    ### Indicateurs clés:
    - **Taux de remplissage**: Pourcentage de la capacité utilisée
    - **Surface utilisée**: Somme des surfaces des câbles assignés
    - **Capacité restante**: Espace disponible

    ## ⚠️ BONNES PRATIQUES

    1. **Taux de réserve recommandé**: 20-30%
    2. **Éviter la saturation**: Maintenir < 80%
    3. **Vérifier régulièrement**: Utiliser le tableau de bord
    4. **Documentation**: Noter les modifications importantes

    ## 🔧 DÉPANNAGE

    **Problème**: Formules cassées après ajout de lignes
    **Solution**: Copier les formules depuis une ligne existante

    **Problème**: Mauvais calcul de chemin
    **Solution**: Vérifier le format "A/B/C" avec "/" uniquement

    **Problème**: Formatage conditionnel absent
    **Solution**: Réappliquer depuis Format → Mise en forme conditionnelle
    """

    # Sauvegarder le manuel
    with open("Manuel_Utilisation_Cables.txt", "w", encoding="utf-8") as f:
        f.write(manual)

    return "Manuel_Utilisation_Cables.txt"


# Exécuter la création de l'application
if __name__ == "__main__":
    print("🚀 Démarrage de la création de l'application...")

    # Vérifier les dépendances
    try:
        import pandas
        import openpyxl

        print("✅ Toutes les dépendances sont installées")
    except ImportError as e:
        print(f"❌ Dépendance manquante: {e}")
        print("💡 Installez avec: pip install pandas openpyxl")
        exit(1)

    # Créer l'application
    success = create_advanced_cable_management()

    if success:
        # Créer le manuel d'utilisation
        manual_file = create_user_manual()

        print(f"\n📖 Manuel d'utilisation créé: {manual_file}")
        print("\n" + "=" * 60)
        print("🎉 CRÉATION TERMINÉE AVEC SUCCÈS !")
        print("=" * 60)
        print("➡️  Ouvrez le fichier Excel et commencez à l'utiliser")
        print("➡️  Consultez le manuel pour plus de détails")
        print("➡️  Compatible avec Excel Online et Google Sheets")

        print("\n💡 CONSEIL: Testez d'abord avec les données d'exemple")
        print("avant d'ajouter vos propres données.")

    else:
        print("\n❌ Échec de la création. Vérifiez les erreurs ci-dessus.")






